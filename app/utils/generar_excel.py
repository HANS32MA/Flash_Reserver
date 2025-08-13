from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import make_response
from io import BytesIO

def generar_excel_usuarios(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios - Flash Reserver"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="28A745")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = ["ID", "Nombre", "Email", "Teléfono", "Rol"]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

    # Filas de datos
    for u in usuarios:
        fila = [
            u.Id,
            u.Nombre,
            u.Email,
            u.Telefono if u.Telefono else "No especificado",
            u.rol.Nombre.title()
        ]
        ws.append(fila)

    # Aplicar bordes y alineación a las celdas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # Ajuste de ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Respuesta HTTP
    response = make_response(output.read())
    response.headers['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
