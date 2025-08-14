from urllib import response
from flask import render_template, make_response
from xhtml2pdf import pisa
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import func
import os
from app import db
from app.models import Usuario, Cancha, Reserva, Categoria, TipoCancha, Rol
from app.auth.decorators import admin_required
from sqlalchemy import func, extract
from app.utils.pdf_utils import generar_pdf
from app.utils.generar_excel import generar_excel_usuarios
from app.models.mensaje import Mensaje
from app.utils.filters import format_currency
from app.models import Post, Like, ComentarioForo




admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ⛔️ Prevención de caché en todas las respuestas del panel admin
@admin_bp.after_request
def agregar_cabeceras_no_cache(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# ---------------------- DASHBOARD ----------------------
@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    # Métricas principales
    total_canchas = Cancha.query.count()
    total_reservas = Reserva.query.count()
    total_usuarios = Usuario.query.count()
    reservas_hoy = Reserva.query.filter(Reserva.Fecha == datetime.now().date()).count()

    # Estadísticas del foro
    total_posts_foro = Post.query.count()
    total_comentarios_foro = ComentarioForo.query.count()
    total_likes_foro = Like.query.count()
    
    # Posts pendientes (ocultos + eliminados)
    posts_ocultos = Post.query.filter_by(Estado='Oculto').count()
    posts_eliminados = Post.query.filter_by(Estado='Eliminado').count()
    posts_pendientes_foro = posts_ocultos + posts_eliminados

    # Últimas 5 reservas
    ultimas_reservas = Reserva.query.order_by(Reserva.FechaCreacion.desc()).limit(5).all()

    # 🔹 Cálculo de canchas populares con reservas, ingresos y ocupación
    canchas_populares = db.session.query(
        Cancha.Nombre.label('Nombre'),
        func.count(Reserva.Id).label('total_reservas'),
        func.sum(Cancha.PrecioHora).label('ingresos'),
        (func.count(Reserva.Id) * 100 / func.sum(func.count(Reserva.Id)).over()).label('ocupacion')
    ).join(Reserva).group_by(Cancha.Id).order_by(
        func.count(Reserva.Id).desc()
    ).limit(5).all()

    # Actividad reciente (ejemplo: últimas 5 reservas confirmadas)
    actividades = db.session.query(
        Reserva.Id.label('id'),
        Usuario.Nombre.label('usuario'),
        Cancha.Nombre.label('cancha'),
        Reserva.Estado.label('estado'),
        Reserva.FechaCreacion.label('fecha')
    ).join(Usuario).join(Cancha).order_by(Reserva.FechaCreacion.desc()).limit(5).all()

    # Ingresos totales
    ingresos_totales = db.session.query(
        func.coalesce(func.sum(Cancha.PrecioHora), 0)
    ).join(Reserva).scalar()

    response = make_response(render_template(
    'admin/dashboard.html',
    total_canchas=total_canchas,
    total_reservas=total_reservas,
    total_usuarios=total_usuarios,
    reservas_hoy=reservas_hoy,
    ultimas_reservas=ultimas_reservas,
    canchas_populares=canchas_populares,
    actividades=actividades,
    ingresos_totales=ingresos_totales,
    format_currency=format_currency,
    total_posts_foro=total_posts_foro,
    total_comentarios_foro=total_comentarios_foro,
    total_likes_foro=total_likes_foro,
    posts_pendientes_foro=posts_pendientes_foro
))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@admin_bp.route('/reportes')
@login_required
@admin_required
def reportes():
    response = make_response(render_template('admin/reportes.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response



@admin_bp.route('/configuracion')
@login_required
@admin_required
def configuracion():
    response = make_response(render_template('admin/configuracion.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response







# ---------------------- CRUD CANCHAS ----------------------
@admin_bp.route('/canchas')
@login_required
@admin_required
def canchas():
    nombre_filter = request.args.get('nombre', '', type=str)
    categoria_filter = request.args.get('categoria', type=int)
    tipo_filter = request.args.get('tipo', type=int)


    query = Cancha.query

    if nombre_filter:
        query = query.filter(Cancha.Nombre.ilike(f"%{nombre_filter}%"))
    if categoria_filter:
        query = query.filter(Cancha.CategoriaId == categoria_filter)
    if tipo_filter:
        query = query.filter(Cancha.TipoCanchaId == tipo_filter)


    canchas = query.all()
    categorias = Categoria.query.all()
    tipos_cancha = TipoCancha.query.all()

    response = make_response(render_template(
        'admin/canchas.html',
        canchas=canchas,
        categorias=categorias,
        tipos_cancha=tipos_cancha,
        nombre_filter=nombre_filter,
        categoria_filter=categoria_filter,
        tipo_filter=tipo_filter,
        format_currency=format_currency
    ))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@admin_bp.route('/cancha/nueva', methods=['GET', 'POST'])
@login_required
@admin_required
def nueva_cancha():
    if request.method == 'POST':
        try:
            cancha = Cancha(
                Nombre=request.form['nombre'],
                Descripcion=request.form.get('descripcion', ''),
                PrecioHora=float(request.form['precio_hora']),
                TipoCanchaId=int(request.form['tipo_cancha']),
                CategoriaId=int(request.form['categoria']),
                Estado='Disponible',
                # Campos de ubicación
                Latitud=float(request.form.get('latitud', 0)) if request.form.get('latitud') else None,
                Longitud=float(request.form.get('longitud', 0)) if request.form.get('longitud') else None,
                Direccion=request.form.get('direccion', '').strip() or None,
                Barrio=request.form.get('barrio', '').strip() or None,
                Ciudad=request.form.get('ciudad', 'Valledupar').strip() or 'Valledupar'
            )

            imagen = request.files.get('imagen')
            if imagen and imagen.filename != '':
                filename = secure_filename(imagen.filename)
                ruta_directorio = os.path.join(current_app.root_path, 'static', 'uploads', 'canchas')
                os.makedirs(ruta_directorio, exist_ok=True)
                ruta_guardado = os.path.join(ruta_directorio, filename)
                imagen.save(ruta_guardado)
                cancha.Imagen = f'uploads/canchas/{filename}'

            db.session.add(cancha)
            db.session.commit()
            flash('Cancha creada exitosamente', 'success')
            return redirect(url_for('admin.canchas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear cancha: {str(e)}', 'error')

    tipos_cancha = TipoCancha.query.all()
    categorias = Categoria.query.all()
    response = make_response(render_template('admin/nueva_cancha.html', tipos=tipos_cancha, categorias=categorias))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@admin_bp.route('/cancha/editar/<int:cancha_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_cancha(cancha_id):
    cancha = Cancha.query.get_or_404(cancha_id)

    if request.method == 'POST':
        try:
            cancha.Nombre = request.form['nombre']
            cancha.Descripcion = request.form.get('descripcion', '').strip() or None
            cancha.PrecioHora = float(request.form['precio_hora'])
            cancha.TipoCanchaId = int(request.form['tipo_cancha'])
            cancha.CategoriaId = int(request.form['categoria'])
            cancha.Estado = request.form.get('estado', cancha.Estado)
            
            # Actualizar campos de ubicación
            cancha.Latitud = float(request.form.get('latitud', 0)) if request.form.get('latitud') else None
            cancha.Longitud = float(request.form.get('longitud', 0)) if request.form.get('longitud') else None
            cancha.Direccion = request.form.get('direccion', '').strip() or None
            cancha.Barrio = request.form.get('barrio', '').strip() or None
            cancha.Ciudad = request.form.get('ciudad', 'Valledupar').strip() or 'Valledupar'

            # Eliminar imagen
            if request.form.get('eliminar_imagen') and cancha.Imagen:
                ruta_anterior = os.path.join(current_app.root_path, 'static', cancha.Imagen)
                if os.path.exists(ruta_anterior):
                    os.remove(ruta_anterior)
                cancha.Imagen = None

            # Nueva imagen
            imagen = request.files.get('imagen')
            if imagen and imagen.filename != '':
                if cancha.Imagen:
                    ruta_anterior = os.path.join(current_app.root_path, 'static', cancha.Imagen)
                    if os.path.exists(ruta_anterior):
                        os.remove(ruta_anterior)

                filename = secure_filename(imagen.filename)
                ruta_directorio = os.path.join(current_app.root_path, 'static', 'uploads', 'canchas')
                os.makedirs(ruta_directorio, exist_ok=True)
                ruta_guardado = os.path.join(ruta_directorio, filename)
                imagen.save(ruta_guardado)
                cancha.Imagen = f'uploads/canchas/{filename}'

            db.session.commit()
            flash('Cancha actualizada exitosamente', 'success')
            return redirect(url_for('admin.canchas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar cancha: {str(e)}', 'error')

    tipos = TipoCancha.query.all()
    categorias = Categoria.query.all()
    response = make_response(render_template('admin/editar_cancha.html', cancha=cancha, tipos=tipos, categorias=categorias))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@admin_bp.route('/cancha/eliminar/<int:cancha_id>', methods=['POST'])
@login_required
@admin_required
def eliminar_cancha(cancha_id):
    cancha = Cancha.query.get_or_404(cancha_id)
    
    # Verificar si hay reservas asociadas
    from app.models.reserva import Reserva
    reservas_activas = Reserva.query.filter_by(CanchaId=cancha_id, Estado='Confirmada').count()
    total_reservas = Reserva.query.filter_by(CanchaId=cancha_id).count()
    
    if total_reservas > 0:
        flash(f'No se puede eliminar la cancha porque tiene {total_reservas} reserva(s) asociada(s) ({reservas_activas} activa(s)). Primero debe cancelar o completar todas las reservas.', 'error')
        return redirect(url_for('admin.canchas'))
    
    try:
        # Eliminar imagen si existe
        if cancha.Imagen:
            ruta_imagen = os.path.join(current_app.root_path, 'static', cancha.Imagen)
            if os.path.exists(ruta_imagen):
                os.remove(ruta_imagen)

        # Eliminar comentarios asociados
        from app.models.comentario import Comentario
        Comentario.query.filter_by(CanchaId=cancha_id).delete()
        
        # Eliminar horarios asociados
        from app.models.horario import Horario
        Horario.query.filter_by(CanchaId=cancha_id).delete()
        
        # Eliminar imágenes asociadas
        from app.models.imagen import Imagen
        Imagen.query.filter_by(CanchaId=cancha_id).delete()
        
        # Ahora eliminar la cancha
        db.session.delete(cancha)
        db.session.commit()
        flash('Cancha eliminada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar cancha: {str(e)}', 'error')

    return redirect(url_for('admin.canchas'))

# ---------------------- CRUD RESERVAS ----------------------
@admin_bp.route('/reservas')
@login_required
@admin_required
def reservas():
    cliente_filter = request.args.get('cliente', '', type=str)
    cancha_filter = request.args.get('cancha', '', type=str)
    fecha_filter = request.args.get('fecha', '', type=str)

    query = Reserva.query.join(Usuario).join(Cancha)

    if cliente_filter:
        query = query.filter(Usuario.Nombre.ilike(f"%{cliente_filter}%"))
    if cancha_filter:
        query = query.filter(Cancha.Nombre == cancha_filter)
    if fecha_filter:
        try:
            fecha_obj = datetime.strptime(fecha_filter, "%Y-%m-%d").date()
            query = query.filter(Reserva.Fecha == fecha_obj)
        except ValueError:
            pass

    reservas = query.order_by(Reserva.FechaCreacion.desc()).all()
    canchas = [c.Nombre for c in Cancha.query.order_by(Cancha.Nombre).all()]

    # Debug: imprimir información de las reservas y usuarios
    print(f"DEBUG ADMIN RESERVAS - Total reservas: {len(reservas)}")
    if reservas:
        for i, reserva in enumerate(reservas[:3]):  # Solo las primeras 3 para debug
            print(f"DEBUG ADMIN RESERVAS - Reserva {i+1}:")
            print(f"  - ID: {reserva.Id}")
            print(f"  - Usuario ID: {reserva.usuario.Id}")
            print(f"  - Usuario Nombre: {reserva.usuario.Nombre}")
            print(f"  - Usuario FotoPerfil: {reserva.usuario.FotoPerfil}")
            print(f"  - Usuario Email: {reserva.usuario.Email}")
            print(f"  - Cancha: {reserva.cancha.Nombre}")
            print(f"  - Fecha: {reserva.Fecha}")

    response = make_response(render_template(
        'admin/reservas.html',
        reservas=reservas,
        canchas=canchas,
        cliente_filter=cliente_filter,
        cancha_filter=cancha_filter,
        fecha_filter=fecha_filter,
        format_currency=format_currency
    ))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@admin_bp.route('/reserva/cancelar/<int:reserva_id>', methods=['POST'])
@login_required
@admin_required
def cancelar_reserva(reserva_id):
    reserva = Reserva.query.get_or_404(reserva_id)
    try:
        # Cambiar estado a cancelada
        reserva.Estado = 'Cancelada'
        db.session.commit()
        
        # Enviar notificación de cancelación por email
        try:
            from app.services.notificacion_service import notificacion_service
            
            # Crear notificación de cancelación
            notificacion_service.notificar_cancelacion_reserva(reserva.Id)
            
            # Cancelar recordatorios automáticos
            from app.services.recordatorio_service import recordatorio_service
            recordatorio_service.cancelar_recordatorios_reserva(reserva.Id)
            
            flash('Reserva cancelada exitosamente y cliente notificado por email', 'success')
            
        except Exception as e:
            # Si falla la notificación, no fallar la cancelación
            flash(f'Reserva cancelada exitosamente, pero error al enviar notificación: {str(e)}', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cancelar reserva: {str(e)}', 'error')

    response = make_response(redirect(url_for('admin.reservas')))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@admin_bp.route('/reserva/confirmar/<int:reserva_id>', methods=['POST'])
@login_required
@admin_required
def confirmar_reserva(reserva_id):
    reserva = Reserva.query.get_or_404(reserva_id)
    try:
        # Cambiar estado a confirmada
        reserva.Estado = 'Confirmada'
        db.session.commit()
        
        # Enviar notificación de confirmación por email
        try:
            from app.services.notificacion_service import notificacion_service
            
            # Crear notificación de confirmación
            notificacion_service.notificar_reserva_confirmada(reserva.Id)
            
            # Programar recordatorios automáticos
            from app.services.recordatorio_service import recordatorio_service
            recordatorio_service.programar_recordatorios_reserva(reserva.Id)
            
            flash('Reserva confirmada exitosamente y cliente notificado por email', 'success')
            
        except Exception as e:
            # Si falla la notificación, no fallar la confirmación
            flash(f'Reserva confirmada exitosamente, pero error al enviar notificación: {str(e)}', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al confirmar reserva: {str(e)}', 'error')

    response = make_response(redirect(url_for('admin.reservas')))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@admin_bp.route('/reserva/<int:reserva_id>')
@login_required
@admin_required
def ver_reserva(reserva_id):
    reserva = Reserva.query.get_or_404(reserva_id)
    return render_template('admin/ver_reserva.html', reserva=reserva, format_currency=format_currency)

# ---------------------- CRUD USUARIOS ----------------------
@admin_bp.route('/usuarios')
@login_required
@admin_required
def usuarios():
    nombre_filter = request.args.get('nombre', '', type=str)
    email_filter = request.args.get('email', '', type=str)
    estado_filter = request.args.get('estado', '', type=str)

    query = Usuario.query.join(Rol)
    if nombre_filter:
        query = query.filter(Usuario.Nombre.ilike(f"%{nombre_filter}%"))
    if email_filter:
        query = query.filter(Usuario.Email.ilike(f"%{email_filter}%"))
    if estado_filter == 'activo':
        query = query.filter(Usuario.Estado == True)
    elif estado_filter == 'inactivo':
        query = query.filter(Usuario.Estado == False)

    usuarios = query.order_by(Usuario.FechaRegistro.desc()).all()

    response = make_response(render_template(
        'admin/usuarios.html',
        usuarios=usuarios,
        nombre_filter=nombre_filter,
        email_filter=email_filter,
        estado_filter=estado_filter
    ))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@admin_bp.route('/usuario/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def nuevo_usuario():
    roles = Rol.query.all()

    if request.method == 'POST':
        try:
            if Usuario.query.filter_by(Email=request.form['email']).first():
                flash('El email ya está registrado', 'error')
                return render_template('admin/nuevo_usuario.html', roles=roles)

            usuario = Usuario(
                Nombre=request.form['nombre_completo'],
                Email=request.form['email'],
                Telefono=request.form['telefono'],
                RolId=int(request.form['rol'])
            )
            usuario.Contrasena = generate_password_hash(request.form['password'])

            db.session.add(usuario)
            db.session.commit()
            flash('Usuario creado exitosamente', 'success')
            return redirect(url_for('admin.usuarios'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear usuario: {str(e)}', 'error')

    response = make_response(render_template('admin/nuevo_usuario.html', roles=roles))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@admin_bp.route('/usuario/editar/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    roles = Rol.query.all()

    if request.method == 'POST':
        try:
            email_existente = Usuario.query.filter_by(Email=request.form['email']).first()
            if email_existente and email_existente.Id != usuario_id:
                flash('El email ya está registrado', 'error')
                return render_template('admin/editar_usuario.html', usuario=usuario, roles=roles)

            usuario.Nombre = request.form['nombre_completo']
            usuario.Email = request.form['email']
            usuario.Telefono = request.form['telefono']
            usuario.RolId = int(request.form['rol'])

            # Manejar el estado del usuario (solo si no es el usuario actual)
            if usuario.Id != current_user.Id:
                usuario.Estado = 'estado' in request.form

            if request.form.get('password'):
                usuario.Contrasena = generate_password_hash(request.form['password'])

            db.session.commit()
            flash('Usuario actualizado exitosamente', 'success')
            return redirect(url_for('admin.usuarios'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar usuario: {str(e)}', 'error')

    response = make_response(render_template('admin/editar_usuario.html', usuario=usuario, roles=roles))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@admin_bp.route('/usuario/eliminar/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def eliminar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    if usuario.Id == current_user.Id:
        flash('No puedes eliminar tu propia cuenta', 'error')
        return redirect(url_for('admin.usuarios'))

    try:
        db.session.delete(usuario)
        db.session.commit()
        flash('Usuario eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar usuario: {str(e)}', 'error')

    return redirect(url_for('admin.usuarios'))

@admin_bp.route('/usuario/cambiar-estado/<int:usuario_id>', methods=['POST'])
@login_required
@admin_required
def cambiar_estado_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # No permitir cambiar el estado de tu propia cuenta
    if usuario.Id == current_user.Id:
        flash('No puedes cambiar el estado de tu propia cuenta', 'error')
        return redirect(url_for('admin.usuarios'))
    
    try:
        # Cambiar el estado (True -> False, False -> True)
        usuario.Estado = not usuario.Estado
        db.session.commit()
        
        estado_texto = "activado" if usuario.Estado else "desactivado"
        flash(f'Usuario {estado_texto} exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar estado del usuario: {str(e)}', 'error')

    return redirect(url_for('admin.usuarios'))

@admin_bp.route('/usuario/<int:usuario_id>', endpoint='ver_usuario')
@login_required
@admin_required
def ver_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    reservas_usuario = Reserva.query.filter_by(
        UsuarioId=usuario_id
    ).order_by(Reserva.FechaCreacion.desc()).all()
    return render_template('admin/ver_usuario.html', usuario=usuario, reservas=reservas_usuario, format_currency=format_currency)

@admin_bp.route('/usuarios/exportar/pdf')
@login_required
@admin_required
def exportar_usuarios_pdf():
    from datetime import datetime
    import os
    usuarios = Usuario.query.all()
    total_usuarios = len(usuarios)
    usuarios_activos = sum(1 for u in usuarios if u.Estado)
    fecha_generacion = datetime.now().strftime('%d/%m/%Y')
    # Ruta absoluta para el logo
    logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
    # Renderizamos la plantilla de usuarios para PDF
    html = render_template(
        'admin/usuarios_pdf.html',
        usuarios=usuarios,
        total_usuarios=total_usuarios,
        usuarios_=usuarios_activos,
        fecha_generacion=fecha_generacion,
        logo_path=logo_path
    )
    return generar_pdf(html, filename="usuarios_registrados.pdf")

@admin_bp.route('/admin/usuarios/excel')
@login_required
@admin_required
def exportar_usuarios_excel():
    usuarios = Usuario.query.all()
    return generar_excel_usuarios(usuarios)

# ---------------------- CRUD CATEGORIAS ----------------------
@admin_bp.route('/categorias')
@login_required
@admin_required
def categorias():
    categorias = Categoria.query.all()
    return render_template('admin/categorias.html', categorias=categorias)

@admin_bp.route('/categoria/nueva', methods=['GET', 'POST'])
@login_required
@admin_required
def nueva_categoria():
    if request.method == 'POST':
        try:
            categoria = Categoria(
                Nombre=request.form['nombre'],
                Descripcion=request.form['descripcion']
            )
            db.session.add(categoria)
            db.session.commit()
            flash('Categoría creada exitosamente', 'success')
            return redirect(url_for('admin.categorias'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear categoría: {str(e)}', 'error')

    return render_template('admin/nueva_categoria.html')

@admin_bp.route('/categoria/editar/<int:categoria_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_categoria(categoria_id):
    categoria = Categoria.query.get_or_404(categoria_id)

    if request.method == 'POST':
        try:
            categoria.Nombre = request.form['nombre']
            categoria.Descripcion = request.form['descripcion']
            db.session.commit()
            flash('Categoría actualizada exitosamente', 'success')
            return redirect(url_for('admin.categorias'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar categoría: {str(e)}', 'error')

    return render_template('admin/editar_categoria.html', categoria=categoria)

@admin_bp.route('/categoria/eliminar/<int:categoria_id>', methods=['POST'])
@login_required
@admin_required
def eliminar_categoria(categoria_id):
    categoria = Categoria.query.get_or_404(categoria_id)

    # Verificar si hay canchas asociadas
    canchas_con_categoria = Cancha.query.filter_by(CategoriaId=categoria_id).count()
    if canchas_con_categoria > 0:
        flash(f'No se puede eliminar la categoría porque hay {canchas_con_categoria} cancha(s) asociada(s)', 'error')
        return redirect(url_for('admin.categorias'))

    try:
        db.session.delete(categoria)
        db.session.commit()
        flash('Categoría eliminada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar categoría: {str(e)}', 'error')

    return redirect(url_for('admin.categorias'))

@admin_bp.route('/categorias/exportar/pdf')
@login_required
@admin_required
def exportar_categorias_pdf():
    from datetime import datetime
    import os
    categorias = Categoria.query.all()
    total_categorias = len(categorias)
    categorias_con_canchas = sum(1 for c in categorias if c.canchas and len(c.canchas) > 0)
    fecha_generacion = datetime.now().strftime('%d/%m/%Y')
    # Ruta absoluta para el logo
    logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
    # Renderizamos la plantilla de categorías para PDF
    html = render_template(
        'admin/categorias_pdf.html',
        categorias=categorias,
        total_categorias=total_categorias,
        categorias_con_canchas=categorias_con_canchas,
        fecha_generacion=fecha_generacion,
        logo_path=logo_path
    )
    return generar_pdf(html, filename="categorias_registradas.pdf")

@admin_bp.route('/reportes/exportar/pdf')
@login_required
@admin_required
def exportar_reportes_pdf():
    from datetime import datetime
    import os
    from sqlalchemy import extract
    
    # Obtener todos los tipos de filtros
    filtro_rapido = request.args.get('filtro_rapido')
    filtro_dia = request.args.get('filtro_dia')
    filtro_mes = request.args.get('filtro_mes')
    filtro_año = request.args.get('filtro_año')
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    # Usar EXACTAMENTE la misma lógica que la API de estadísticas
    hoy = datetime.now().date()
    
    query_reservas = Reserva.query
    
    # Aplicar filtros según prioridad (misma lógica que la API)
    if filtro_rapido:
        # Filtros rápidos
        if filtro_rapido == 'hoy':
            query_reservas = query_reservas.filter(Reserva.Fecha == hoy)
        elif filtro_rapido == 'ayer':
            ayer = hoy - timedelta(days=1)
            query_reservas = query_reservas.filter(Reserva.Fecha == ayer)
        elif filtro_rapido == 'semana':
            inicio_semana = hoy - timedelta(days=hoy.weekday())
            fin_semana = inicio_semana + timedelta(days=6)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_semana,
                Reserva.Fecha <= fin_semana
            )
        elif filtro_rapido == 'mes':
            inicio_mes = hoy.replace(day=1)
            fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_mes,
                Reserva.Fecha <= fin_mes
            )
        elif filtro_rapido == 'trimestre':
            trimestre = (hoy.month - 1) // 3
            inicio_trimestre = hoy.replace(month=trimestre * 3 + 1, day=1)
            fin_trimestre = (inicio_trimestre + timedelta(days=93)).replace(day=1) - timedelta(days=1)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_trimestre,
                Reserva.Fecha <= fin_trimestre
            )
        elif filtro_rapido == 'año':
            inicio_año = hoy.replace(month=1, day=1)
            fin_año = hoy.replace(month=12, day=31)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_año,
                Reserva.Fecha <= fin_año
            )
    elif filtro_dia:
        # Filtro por día específico
        query_reservas = query_reservas.filter(Reserva.Fecha == filtro_dia)
    elif filtro_mes:
        # Filtro por mes específico
        año, mes = filtro_mes.split('-')
        query_reservas = query_reservas.filter(
            extract('year', Reserva.Fecha) == int(año),
            extract('month', Reserva.Fecha) == int(mes)
        )
    elif filtro_año:
        # Filtro por año específico
        query_reservas = query_reservas.filter(extract('year', Reserva.Fecha) == int(filtro_año))
    elif fecha_inicio and fecha_fin:
        # Filtro por rango personalizado
        query_reservas = query_reservas.filter(
            Reserva.Fecha >= fecha_inicio,
            Reserva.Fecha <= fecha_fin
        )
    
    # --- Totales ---
    total_usuarios = Usuario.query.count()
    total_canchas = Cancha.query.count()
    total_reservas = query_reservas.count()
    reservas_hoy = (
        db.session.query(func.count(Reserva.Id))
        .filter(func.date(Reserva.FechaCreacion) == hoy)
        .scalar()
    )
    
    # --- Reservas por mes + ingresos (EXACTO como la API) ---
    reservas_por_mes = []
    ingresos_por_mes = []
    meses_labels = []
    for mes in range(1, 13):
        count_mes = query_reservas.filter(extract('month', Reserva.Fecha) == mes).count()
        ingresos_mes = (
            db.session.query(func.coalesce(func.sum(Cancha.PrecioHora), 0))
            .join(Reserva, Reserva.CanchaId == Cancha.Id)
            .filter(extract('month', Reserva.Fecha) == mes)
        )
        if fecha_inicio and fecha_fin:
            ingresos_mes = ingresos_mes.filter(
                Reserva.Fecha >= fecha_inicio,
                Reserva.Fecha <= fecha_fin
            )
        ingresos_mes = ingresos_mes.scalar() or 0

        reservas_por_mes.append(count_mes)
        ingresos_por_mes.append(float(ingresos_mes))
        meses_labels.append(datetime(2025, mes, 1).strftime('%b'))  # Usar %b como la API
    
    # --- Ocupación por cancha (EXACTO como la API) ---
    canchas_populares_query = (
        db.session.query(
            Cancha.Nombre.label('nombre'),
            func.count(Reserva.Id).label('total_reservas'),
            func.coalesce(func.sum(Cancha.PrecioHora), 0).label('ingresos')
        )
        .join(Reserva, Cancha.Id == Reserva.CanchaId)
        .group_by(Cancha.Id)
        .order_by(func.count(Reserva.Id).desc())
        .limit(5)
        .all()
    )
    
    ocupacion_labels = []
    ocupacion_values = []
    for c in canchas_populares_query:
        ocupacion_labels.append(c.nombre)
        ocupacion_porcentaje = round((c.total_reservas / max(1, total_reservas)) * 100, 2)
        ocupacion_values.append(ocupacion_porcentaje)
    
    # --- Heatmap semanal (EXACTO como la API) ---
    reservas_semana = Reserva.query.filter(
        Reserva.Fecha >= hoy - timedelta(days=6),
        Reserva.Fecha <= hoy
    ).all()

    heatmap = []
    for day in range(7):  
        for hour in range(24):
            heatmap.append({"x": hour, "y": day, "v": 0})
    for r in reservas_semana:
        day_index = r.Fecha.weekday()  # lunes=0
        hour_index = r.HoraInicio.hour
        idx = day_index * 24 + hour_index
        heatmap[idx]["v"] += 1
    
    fecha_generacion = datetime.now().strftime('%d/%m/%Y')
    logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
    
    html = render_template(
        'admin/reportes_pdf.html',
        total_usuarios=total_usuarios,
        total_canchas=total_canchas,
        total_reservas=total_reservas,
        reservas_hoy=reservas_hoy,
        meses=meses_labels,
        reservas_por_mes=reservas_por_mes,
        ingresos_por_mes=ingresos_por_mes,
        ocupacion_labels=ocupacion_labels,
        ocupacion_values=ocupacion_values,
        heatmap=heatmap,
        fecha_generacion=fecha_generacion,
        logo_path=logo_path,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )
    return generar_pdf(html, filename="reportes_generales.pdf")

@admin_bp.route('/reportes/exportar/excel')
@login_required
@admin_required
def exportar_reportes_excel():
    from datetime import datetime
    import os
    from sqlalchemy import extract
    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.worksheet.worksheet import Worksheet
    
    # Obtener todos los tipos de filtros
    filtro_rapido = request.args.get('filtro_rapido')
    filtro_dia = request.args.get('filtro_dia')
    filtro_mes = request.args.get('filtro_mes')
    filtro_año = request.args.get('filtro_año')
    fecha_inicio = request.args.get('inicio')
    fecha_fin = request.args.get('fin')
    
    # Usar EXACTAMENTE la misma lógica que la API de estadísticas
    hoy = datetime.now().date()
    
    query_reservas = Reserva.query
    
    # Aplicar filtros según prioridad (misma lógica que la API)
    if filtro_rapido:
        # Filtros rápidos
        if filtro_rapido == 'hoy':
            query_reservas = query_reservas.filter(Reserva.Fecha == hoy)
        elif filtro_rapido == 'ayer':
            ayer = hoy - timedelta(days=1)
            query_reservas = query_reservas.filter(Reserva.Fecha == ayer)
        elif filtro_rapido == 'semana':
            inicio_semana = hoy - timedelta(days=hoy.weekday())
            fin_semana = inicio_semana + timedelta(days=6)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_semana,
                Reserva.Fecha <= fin_semana
            )
        elif filtro_rapido == 'mes':
            inicio_mes = hoy.replace(day=1)
            fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_mes,
                Reserva.Fecha <= fin_mes
            )
        elif filtro_rapido == 'trimestre':
            trimestre = (hoy.month - 1) // 3
            inicio_trimestre = hoy.replace(month=trimestre * 3 + 1, day=1)
            fin_trimestre = (inicio_trimestre + timedelta(days=93)).replace(day=1) - timedelta(days=1)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_trimestre,
                Reserva.Fecha <= fin_trimestre
            )
        elif filtro_rapido == 'año':
            inicio_año = hoy.replace(month=1, day=1)
            fin_año = hoy.replace(month=12, day=31)
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio_año,
                Reserva.Fecha <= fin_año
            )
    elif filtro_dia:
        # Filtro por día específico
        query_reservas = query_reservas.filter(Reserva.Fecha == filtro_dia)
    elif filtro_mes:
        # Filtro por mes específico
        año, mes = filtro_mes.split('-')
        query_reservas = query_reservas.filter(
            extract('year', Reserva.Fecha) == int(año),
            extract('month', Reserva.Fecha) == int(mes)
        )
    elif filtro_año:
        # Filtro por año específico
        query_reservas = query_reservas.filter(extract('year', Reserva.Fecha) == int(filtro_año))
    elif fecha_inicio and fecha_fin:
        # Filtro por rango personalizado
        query_reservas = query_reservas.filter(
            Reserva.Fecha >= fecha_inicio,
            Reserva.Fecha <= fecha_fin
        )
    
    # Obtener datos para Excel
    reservas_data = []
    for reserva in query_reservas.join(Usuario).join(Cancha).all():
        reservas_data.append({
            'ID': reserva.Id,
            'Usuario': reserva.usuario.Nombre,
            'Cancha': reserva.cancha.Nombre,
            'Fecha': reserva.Fecha.strftime('%d/%m/%Y'),
            'Hora Inicio': reserva.HoraInicio.strftime('%H:%M'),
            'Hora Fin': reserva.HoraFin.strftime('%H:%M'),
            'Duración (horas)': reserva.duracion_horas,
            'Estado': reserva.Estado,
            'Precio Total': float(reserva.precio_total),
            'Fecha Creación': reserva.FechaCreacion.strftime('%d/%m/%Y %H:%M')
        })
    
    # Crear archivo Excel con openpyxl para mejor control del diseño
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # ==================== HOJA DE RESERVAS ====================
    ws_reservas = wb.create_sheet("Reservas")
    
    # Definir estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'ID', 'Usuario', 'Cancha', 'Fecha', 'Hora Inicio', 'Hora Fin',
        'Duración (horas)', 'Estado', 'Precio Total (COP)', 'Fecha Creación'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws_reservas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row, reserva in enumerate(reservas_data, 2):
        for col, (key, value) in enumerate(reserva.items(), 1):
            cell = ws_reservas.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.border = border
            
            # Aplicar formato específico según el tipo de dato
            if key == 'ID':
                cell.alignment = center_alignment
            elif key in ['Hora Inicio', 'Hora Fin']:
                cell.alignment = center_alignment
            elif key == 'Duración (horas)':
                cell.alignment = center_alignment
                cell.number_format = '0.0'
            elif key == 'Estado':
                cell.alignment = center_alignment
                # Aplicar color según el estado
                if value == 'Confirmada':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == 'Pendiente':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif value == 'Cancelada':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif key == 'Precio Total':
                cell.alignment = right_alignment
                cell.number_format = '#,##0'
            elif key == 'Fecha Creación':
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
    
    # Ajustar ancho de columnas
    column_widths = [8, 25, 20, 12, 12, 12, 15, 12, 18, 20]
    for col, width in enumerate(column_widths, 1):
        ws_reservas.column_dimensions[ws_reservas.cell(row=1, column=col).column_letter].width = width
    
    # Añadir título y filtros aplicados
    ws_reservas.insert_rows(1, 3)
    
    # Título principal
    title_cell = ws_reservas.cell(row=1, column=1, value="REPORTE DE RESERVAS - FLASH RESERVER")
    title_cell.font = Font(name='Arial', size=16, bold=True, color='366092')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws_reservas.merge_cells('A1:J1')
    
    # Filtros aplicados
    filtro_texto = "Filtros aplicados: "
    if filtro_rapido:
        filtro_texto += f"Filtro rápido: {filtro_rapido.title()}"
    elif filtro_dia:
        filtro_texto += f"Día: {filtro_dia}"
    elif filtro_mes:
        filtro_texto += f"Mes: {filtro_mes}"
    elif filtro_año:
        filtro_texto += f"Año: {filtro_año}"
    elif fecha_inicio and fecha_fin:
        filtro_texto += f"Rango: {fecha_inicio} a {fecha_fin}"
    else:
        filtro_texto += "Todos los registros"
    
    filtro_cell = ws_reservas.cell(row=2, column=1, value=filtro_texto)
    filtro_cell.font = Font(name='Arial', size=11, italic=True, color='666666')
    filtro_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws_reservas.merge_cells('A2:J2')
    
    # Fecha de generación
    fecha_gen = ws_reservas.cell(row=3, column=1, value=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    fecha_gen.font = Font(name='Arial', size=10, color='666666')
    fecha_gen.alignment = Alignment(horizontal='left', vertical='center')
    ws_reservas.merge_cells('A3:J3')
    
    # Ajustar encabezados de datos (ahora en fila 4)
    for col in range(1, len(headers) + 1):
        cell = ws_reservas.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # ==================== HOJA DE ESTADÍSTICAS ====================
    ws_stats = wb.create_sheet("Estadísticas")
    
    # Título
    title_cell = ws_stats.cell(row=1, column=1, value="ESTADÍSTICAS GENERALES")
    title_cell.font = Font(name='Arial', size=16, bold=True, color='366092')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_stats.merge_cells('A1:C1')
    
    # Encabezados de estadísticas
    stats_headers = ['Métrica', 'Valor', 'Descripción']
    for col, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos de estadísticas
    stats_data = [
        ('Total Usuarios', Usuario.query.count(), 'Número total de usuarios registrados'),
        ('Total Canchas', Cancha.query.count(), 'Número total de canchas disponibles'),
        ('Total Reservas', len(reservas_data), 'Número total de reservas en el período'),
        ('Reservas Hoy', Reserva.query.filter(Reserva.Fecha == hoy).count(), 'Reservas realizadas hoy'),
        ('Ingresos Totales (COP)', sum(r['Precio Total'] for r in reservas_data), 'Ingresos totales en el período'),
        ('Promedio por Reserva', round(sum(r['Precio Total'] for r in reservas_data) / len(reservas_data) if reservas_data else 0, 0), 'Precio promedio por reserva'),
        ('Cancha Más Popular', '', 'Cancha con más reservas'),
        ('Usuario Más Activo', '', 'Usuario con más reservas')
    ]
    
    # Obtener cancha más popular
    if reservas_data:
        cancha_counts = {}
        for r in reservas_data:
            cancha_counts[r['Cancha']] = cancha_counts.get(r['Cancha'], 0) + 1
        cancha_popular = max(cancha_counts, key=cancha_counts.get)
        stats_data[6] = (stats_data[6][0], cancha_popular, stats_data[6][2])
    
    # Obtener usuario más activo
    if reservas_data:
        usuario_counts = {}
        for r in reservas_data:
            usuario_counts[r['Usuario']] = usuario_counts.get(r['Usuario'], 0) + 1
        usuario_activo = max(usuario_counts, key=usuario_counts.get)
        stats_data[7] = (stats_data[7][0], usuario_activo, stats_data[7][2])
    
    # Escribir estadísticas
    for row, (metrica, valor, descripcion) in enumerate(stats_data, 4):
        # Métrica
        cell = ws_stats.cell(row=row, column=1, value=metrica)
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.border = border
        cell.alignment = data_alignment
        
        # Valor
        cell = ws_stats.cell(row=row, column=2, value=valor)
        cell.font = data_font
        cell.border = border
        if 'Precio' in metrica or 'Ingresos' in metrica:
            cell.number_format = '#,##0'
            cell.alignment = right_alignment
        else:
            cell.alignment = center_alignment
        
        # Descripción
        cell = ws_stats.cell(row=row, column=3, value=descripcion)
        cell.font = Font(name='Arial', size=10, italic=True, color='666666')
        cell.border = border
        cell.alignment = data_alignment
    
    # Ajustar ancho de columnas
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20
    ws_stats.column_dimensions['C'].width = 50
    
    # ==================== HOJA DE RESUMEN ====================
    ws_resumen = wb.create_sheet("Resumen")
    
    # Título
    title_cell = ws_resumen.cell(row=1, column=1, value="RESUMEN EJECUTIVO")
    title_cell.font = Font(name='Arial', size=16, bold=True, color='366092')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_resumen.merge_cells('A1:D1')
    
    # Información del reporte
    info_data = [
        ('Período del Reporte', filtro_texto.replace('Filtros aplicados: ', '')),
        ('Total de Registros', len(reservas_data)),
        ('Fecha de Generación', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ('Generado por', current_user.Nombre if current_user else 'Sistema')
    ]
    
    for row, (label, value) in enumerate(info_data, 3):
        # Label
        cell = ws_resumen.cell(row=row, column=1, value=label)
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # Valor
        cell = ws_resumen.cell(row=row, column=2, value=value)
        cell.font = data_font
        cell.border = border
        cell.alignment = data_alignment
        
        ws_resumen.merge_cells(f'A{row}:B{row}')
    
    # Ajustar ancho de columnas
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 40
    
    # Guardar archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generar nombre de archivo
    filtro_applied = ""
    if filtro_rapido:
        filtro_applied = f"_{filtro_rapido}"
    elif filtro_dia:
        filtro_applied = f"_dia_{filtro_dia}"
    elif filtro_mes:
        filtro_applied = f"_mes_{filtro_mes}"
    elif filtro_año:
        filtro_applied = f"_año_{filtro_año}"
    elif fecha_inicio and fecha_fin:
        filtro_applied = f"_desde_{fecha_inicio}_hasta_{fecha_fin}"
    
    filename = f"reportes_reservas{filtro_applied}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ---------------------- CRUD TIPOS DE CANCHA ----------------------
@admin_bp.route('/tipos-cancha')
@login_required
@admin_required
def tipos_cancha():
    tipos = TipoCancha.query.all()
    return render_template('admin/tipos_cancha.html', tipos=tipos)

@admin_bp.route('/tipo-cancha/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def nuevo_tipo_cancha():
    if request.method == 'POST':
        try:
            tipo = TipoCancha(
                Nombre=request.form['nombre'],
                Descripcion=request.form['descripcion']
            )
            db.session.add(tipo)
            db.session.commit()
            flash('Tipo de cancha creado exitosamente', 'success')
            return redirect(url_for('admin.tipos_cancha'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear tipo de cancha: {str(e)}', 'error')

    return render_template('admin/nuevo_tipo_cancha.html')

@admin_bp.route('/tipo-cancha/editar/<int:tipo_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_tipo_cancha(tipo_id):
    tipo = TipoCancha.query.get_or_404(tipo_id)

    if request.method == 'POST':
        try:
            tipo.Nombre = request.form['nombre']
            tipo.Descripcion = request.form['descripcion']
            db.session.commit()
            flash('Tipo de cancha actualizado exitosamente', 'success')
            return redirect(url_for('admin.tipos_cancha'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar tipo de cancha: {str(e)}', 'error')

    return render_template('admin/editar_tipo_cancha.html', tipo=tipo)

@admin_bp.route('/tipo-cancha/eliminar/<int:tipo_id>', methods=['POST'])
@login_required
@admin_required
def eliminar_tipo_cancha(tipo_id):
    tipo = TipoCancha.query.get_or_404(tipo_id)
    canchas_con_tipo = Cancha.query.filter_by(TipoCanchaId=tipo_id).count()
    if canchas_con_tipo > 0:
        flash(f'No se puede eliminar el tipo porque hay {canchas_con_tipo} cancha(s) asociada(s)', 'error')
        return redirect(url_for('admin.tipos_cancha'))

    try:
        db.session.delete(tipo)
        db.session.commit()
        flash('Tipo de cancha eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar tipo de cancha: {str(e)}', 'error')

    return redirect(url_for('admin.tipos_cancha'))

# ---------------------- PERFIL ----------------------
@admin_bp.route('/perfil')
@login_required
def perfil():
    response = make_response(render_template('admin/perfil.html', usuario=current_user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response





@admin_bp.route('/perfil/editar', methods=['GET', 'POST'])
@login_required
def editar_perfil():
    usuario = current_user

    if request.method == 'POST':
        usuario.Nombre = request.form['nombre_completo']
        usuario.Email = request.form['email']
        usuario.Telefono = request.form['telefono']

        foto = request.files.get('foto')
        if foto and foto.filename != '':
            if usuario.FotoPerfil and usuario.FotoPerfil != 'img/default-user.png':
                ruta_anterior = os.path.join(current_app.root_path, 'static', usuario.FotoPerfil)
                if os.path.exists(ruta_anterior):
                    os.remove(ruta_anterior)

            filename = secure_filename(foto.filename)
            ruta_directorio = os.path.join(current_app.root_path, 'static', 'uploads', 'perfiles')
            os.makedirs(ruta_directorio, exist_ok=True)
            ruta_guardado = os.path.join(ruta_directorio, filename)
            foto.save(ruta_guardado)
            usuario.FotoPerfil = f'uploads/perfiles/{filename}'

        db.session.commit()
        flash('Perfil actualizado exitosamente', 'success')
        return redirect(url_for('admin.perfil'))

    response = make_response(render_template('admin/editar_perfil.html', usuario=usuario))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ---------------------- API ESTADISTICAS ----------------------
@admin_bp.route('/api/estadisticas')
@login_required
@admin_required
def api_estadisticas():
    try:
        from sqlalchemy import extract

        hoy = datetime.now().date()
        
        # Obtener todos los tipos de filtros
        filtro_rapido = request.args.get('filtro_rapido')
        filtro_dia = request.args.get('filtro_dia')
        filtro_mes = request.args.get('filtro_mes')
        filtro_año = request.args.get('filtro_año')
        inicio = request.args.get('inicio')
        fin = request.args.get('fin')

        query_reservas = Reserva.query
        
        # Aplicar filtros según prioridad
        if filtro_rapido:
            # Filtros rápidos
            if filtro_rapido == 'hoy':
                query_reservas = query_reservas.filter(Reserva.Fecha == hoy)
            elif filtro_rapido == 'ayer':
                ayer = hoy - timedelta(days=1)
                query_reservas = query_reservas.filter(Reserva.Fecha == ayer)
            elif filtro_rapido == 'semana':
                inicio_semana = hoy - timedelta(days=hoy.weekday())
                fin_semana = inicio_semana + timedelta(days=6)
                query_reservas = query_reservas.filter(
                    Reserva.Fecha >= inicio_semana,
                    Reserva.Fecha <= fin_semana
                )
            elif filtro_rapido == 'mes':
                inicio_mes = hoy.replace(day=1)
                fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                query_reservas = query_reservas.filter(
                    Reserva.Fecha >= inicio_mes,
                    Reserva.Fecha <= fin_mes
                )
            elif filtro_rapido == 'trimestre':
                trimestre = (hoy.month - 1) // 3
                inicio_trimestre = hoy.replace(month=trimestre * 3 + 1, day=1)
                fin_trimestre = (inicio_trimestre + timedelta(days=93)).replace(day=1) - timedelta(days=1)
                query_reservas = query_reservas.filter(
                    Reserva.Fecha >= inicio_trimestre,
                    Reserva.Fecha <= fin_trimestre
                )
            elif filtro_rapido == 'año':
                inicio_año = hoy.replace(month=1, day=1)
                fin_año = hoy.replace(month=12, day=31)
                query_reservas = query_reservas.filter(
                    Reserva.Fecha >= inicio_año,
                    Reserva.Fecha <= fin_año
                )
        elif filtro_dia:
            # Filtro por día específico
            query_reservas = query_reservas.filter(Reserva.Fecha == filtro_dia)
        elif filtro_mes:
            # Filtro por mes específico
            año, mes = filtro_mes.split('-')
            query_reservas = query_reservas.filter(
                extract('year', Reserva.Fecha) == int(año),
                extract('month', Reserva.Fecha) == int(mes)
            )
        elif filtro_año:
            # Filtro por año específico
            query_reservas = query_reservas.filter(extract('year', Reserva.Fecha) == int(filtro_año))
        elif inicio and fin:
            # Filtro por rango personalizado
            query_reservas = query_reservas.filter(
                Reserva.Fecha >= inicio,
                Reserva.Fecha <= fin
            )

        # --- Totales ---
        total_usuarios = Usuario.query.count()
        total_canchas = Cancha.query.count()
        total_reservas = query_reservas.count()
        reservas_hoy = (
            db.session.query(func.count(Reserva.Id))
            .filter(func.date(Reserva.FechaCreacion) == hoy)
            .scalar()
        )


        # --- Ingresos totales ---
        ingresos_totales = (
            db.session.query(func.coalesce(func.sum(Cancha.PrecioHora), 0))
            .join(Reserva, Reserva.CanchaId == Cancha.Id)
            .filter(Reserva.Estado == "Confirmada")
        )
        if inicio and fin:
            ingresos_totales = ingresos_totales.filter(
                Reserva.Fecha >= inicio,
                Reserva.Fecha <= fin
            )
        ingresos_totales = ingresos_totales.scalar() or 0

        # --- Últimas reservas ---
        ultimas_reservas = [
            {
                "id": r.Id,
                "usuario": r.usuario.Nombre,
                "cancha": r.cancha.Nombre,
                "fecha": r.Fecha.strftime('%d/%m/%Y %H:%M'),
                "estado": r.Estado
            }
            for r in query_reservas.join(Usuario).join(Cancha)
            .order_by(Reserva.FechaCreacion.desc())
            .limit(5)
        ]

        # --- Canchas más populares ---
        canchas_populares_query = (
            db.session.query(
                Cancha.Nombre.label('nombre'),
                func.count(Reserva.Id).label('total_reservas'),
                func.coalesce(func.sum(Cancha.PrecioHora), 0).label('ingresos')
            )
            .join(Reserva, Cancha.Id == Reserva.CanchaId)
            .group_by(Cancha.Id)
            .order_by(func.count(Reserva.Id).desc())
            .limit(5)
            .all()
        )

        canchas_populares = [
            {
                "nombre": c.nombre,
                "total_reservas": c.total_reservas,
                "ingresos": float(c.ingresos),
                "ocupacion": round((c.total_reservas / max(1, total_reservas)) * 100, 2)
            }
            for c in canchas_populares_query
        ]

        # --- Actividad reciente ---
        actividades = [
            {
                "usuario": a.usuario.Nombre,
                "accion": f"Realizó una reserva en {a.cancha.Nombre}" if a.Estado == "Confirmada" else f"Reserva {a.Estado}",
                "fecha": a.FechaCreacion.strftime('%d/%m/%Y %H:%M')
            }
            for a in query_reservas.join(Usuario)
            .order_by(Reserva.FechaCreacion.desc())
            .limit(5)
        ]

        # --- Reservas por mes + ingresos ---
        reservas_por_mes = []
        ingresos_por_mes = []
        meses_labels = []
        for mes in range(1, 13):
            count_mes = query_reservas.filter(extract('month', Reserva.Fecha) == mes).count()
            ingresos_mes = (
                db.session.query(func.coalesce(func.sum(Cancha.PrecioHora), 0))
                .join(Reserva, Reserva.CanchaId == Cancha.Id)
                .filter(extract('month', Reserva.Fecha) == mes)
            )
            if inicio and fin:
                ingresos_mes = ingresos_mes.filter(
                    Reserva.Fecha >= inicio,
                    Reserva.Fecha <= fin
                )
            ingresos_mes = ingresos_mes.scalar() or 0

            reservas_por_mes.append(count_mes)
            ingresos_por_mes.append(float(ingresos_mes))
            meses_labels.append(datetime(2025, mes, 1).strftime('%b'))

        # --- Ocupación para gráfico de barras ---
        ocupacion_labels = [c['nombre'] for c in canchas_populares]
        ocupacion_values = [c['ocupacion'] for c in canchas_populares]

        # --- Heatmap semanal ---
        reservas_semana = Reserva.query.filter(
            Reserva.Fecha >= hoy - timedelta(days=6),
            Reserva.Fecha <= hoy
        ).all()

        heatmap = []
        for day in range(7):  
            for hour in range(24):
                heatmap.append({"x": hour, "y": day, "v": 0})
        for r in reservas_semana:
            day_index = r.Fecha.weekday()  # lunes=0
            hour_index = r.HoraInicio.hour
            idx = day_index * 24 + hour_index
            heatmap[idx]["v"] += 1

        return jsonify({
            "total_usuarios": total_usuarios,
            "total_canchas": total_canchas,
            "total_reservas": total_reservas,
            "reservas_hoy": reservas_hoy,
            "ingresos_totales": float(ingresos_totales),
            "ultimas_reservas": ultimas_reservas,
            "canchas_populares": canchas_populares,
            "actividades": actividades,
            "meses": meses_labels,
            "reservas_por_mes": reservas_por_mes,
            "ingresos_por_mes": ingresos_por_mes,
            "ocupacion_labels": ocupacion_labels,
            "ocupacion_values": ocupacion_values,
            "heatmap": heatmap
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/notificaciones')
@login_required
@admin_required
def obtener_notificaciones():
    try:
        # Verificar si hay mensajes en la base de datos
        total_mensajes = Mensaje.query.count()
        
        if total_mensajes == 0:
            return jsonify({
                'success': True,
                'notifications': [],
                'unread_count': 0
            })
        
        # Obtener conversaciones agrupadas por usuario
        from sqlalchemy import func
        
        # Subconsulta para obtener el último mensaje de cada usuario
        subquery = db.session.query(
            Mensaje.UsuarioId,
            func.max(Mensaje.FechaEnvio).label('ultima_fecha')
        ).group_by(Mensaje.UsuarioId).subquery()
        
        # Obtener los mensajes más recientes de cada usuario
        mensajes_recientes = db.session.query(Mensaje).join(
            subquery,
            db.and_(
                Mensaje.UsuarioId == subquery.c.UsuarioId,
                Mensaje.FechaEnvio == subquery.c.ultima_fecha
            )
        ).order_by(Mensaje.FechaEnvio.desc()).all()
        
        notifications = []
        for mensaje in mensajes_recientes:
            try:
                # Contar mensajes no leídos de este usuario
                mensajes_no_leidos = Mensaje.query.filter_by(
                    UsuarioId=mensaje.UsuarioId,
                    Leido=False
                ).count()
                
                # Obtener todos los mensajes de este usuario para el conteo
                total_mensajes_usuario = Mensaje.query.filter_by(UsuarioId=mensaje.UsuarioId).count()
                
                # Verificar que el usuario existe
                if not mensaje.usuario:
                    continue
                
                notification = {
                    'id': mensaje.Id,
                    'usuario_id': mensaje.UsuarioId,
                    'asunto': mensaje.Asunto or 'Sin asunto',
                    'mensaje': mensaje.Mensaje or 'Sin mensaje',
                    'fecha': mensaje.fecha_formateada,
                    'leido': mensaje.Leido,
                    'usuario_nombre': f"{mensaje.usuario.Nombre or ''}".strip(),
                    'respuesta': mensaje.Respuesta,
                    'fecha_respuesta': mensaje.fecha_respuesta_formateada if mensaje.Respuesta else None,
                    'mensajes_no_leidos': mensajes_no_leidos,
                    'total_mensajes': total_mensajes_usuario
                }
                notifications.append(notification)
            except Exception as e:
                print(f"Error procesando mensaje {mensaje.Id}: {e}")
                continue
        
        # Contar total de mensajes no leídos
        total_unread = db.session.query(Mensaje).filter_by(Leido=False).count()
        
        return jsonify({
            'success': True,
            'notifications': notifications,
            'unread_count': total_unread
        })
        
    except Exception as e:
        print(f"Error en obtener_notificaciones: {e}")
        return jsonify({
            'success': False, 
            'error': str(e),
            'notifications': [],
            'unread_count': 0
        })

@admin_bp.route('/responder-mensaje', methods=['POST'])
@login_required
@admin_required
def responder_mensaje():
    try:
        data = request.get_json()
        notification_id = data.get('notification_id')
        respuesta = data.get('respuesta')
        
        if not notification_id or not respuesta:
            return jsonify({'success': False, 'error': 'ID de notificación y respuesta son requeridos'})
        
        # Buscar el mensaje
        mensaje = Mensaje.query.get(notification_id)
        if not mensaje:
            return jsonify({'success': False, 'error': 'Mensaje no encontrado'})
        
        # Actualizar con la respuesta
        mensaje.Respuesta = respuesta
        mensaje.FechaRespuesta = datetime.utcnow()
        mensaje.Leido = True
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Respuesta enviada correctamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/marcar-notificaciones-leidas', methods=['POST'])
@login_required
@admin_required
def marcar_notificaciones_leidas():
    try:
        # Marcar todos los mensajes como leídos
        Mensaje.query.filter_by(Leido=False).update({'Leido': True})
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Todas las notificaciones marcadas como leídas'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/eliminar-mensaje', methods=['POST'])
@login_required
@admin_required
def eliminar_mensaje():
    try:
        data = request.get_json()
        mensaje_id = data.get('mensaje_id')
        
        if not mensaje_id:
            return jsonify({'success': False, 'error': 'ID de mensaje es requerido'})
        
        # Buscar y eliminar el mensaje
        mensaje = Mensaje.query.get(mensaje_id)
        if not mensaje:
            return jsonify({'success': False, 'error': 'Mensaje no encontrado'})
        
        db.session.delete(mensaje)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Mensaje eliminado correctamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/marcar-mensaje-leido', methods=['POST'])
@login_required
@admin_required
def marcar_mensaje_leido():
    try:
        data = request.get_json()
        mensaje_id = data.get('mensaje_id')
        
        if not mensaje_id:
            return jsonify({'success': False, 'error': 'ID de mensaje es requerido'})
        
        # Buscar y marcar el mensaje como leído
        mensaje = Mensaje.query.get(mensaje_id)
        if not mensaje:
            return jsonify({'success': False, 'error': 'Mensaje no encontrado'})
        
        mensaje.Leido = True
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Mensaje marcado como leído'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/conversacion/<int:usuario_id>')
@login_required
@admin_required
def obtener_conversacion(usuario_id):
    try:
        # Obtener todos los mensajes del usuario ordenados por fecha
        mensajes = Mensaje.query.filter_by(UsuarioId=usuario_id).order_by(Mensaje.FechaEnvio).all()
        
        conversacion = []
        for mensaje in mensajes:
            conversacion.append({
                'id': mensaje.Id,
                'mensaje': mensaje.Mensaje,
                'fecha': mensaje.fecha_formateada,
                'leido': mensaje.Leido,
                'respuesta': mensaje.Respuesta,
                'fecha_respuesta': mensaje.fecha_respuesta_formateada if mensaje.Respuesta else None
            })
        
        return jsonify({
            'success': True,
            'conversacion': conversacion
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/marcar-conversacion-leida', methods=['POST'])
@login_required
@admin_required
def marcar_conversacion_leida():
    try:
        data = request.get_json()
        usuario_id = data.get('usuario_id')
        
        if not usuario_id:
            return jsonify({'success': False, 'error': 'ID de usuario es requerido'})
        
        # Marcar todos los mensajes del usuario como leídos
        Mensaje.query.filter_by(UsuarioId=usuario_id, Leido=False).update({'Leido': True})
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Conversación marcada como leída'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/eliminar-conversacion', methods=['POST'])
@login_required
@admin_required
def eliminar_conversacion():
    try:
        data = request.get_json()
        usuario_id = data.get('usuario_id')
        
        if not usuario_id:
            return jsonify({'success': False, 'error': 'ID de usuario es requerido'})
        
        # Eliminar todos los mensajes del usuario
        Mensaje.query.filter_by(UsuarioId=usuario_id).delete()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Conversación eliminada correctamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# ---------------------- FIN PAGINA --------------------

# ===== RUTAS DEL FORO PARA ADMINISTRADORES =====

@admin_bp.route('/admin_gestionar_posts')
@login_required
@admin_required
def admin_gestionar_posts():
    """Vista para que el administrador gestione todos los posts del foro"""
    # Debug: Imprimir información del usuario actual
    print(f"Usuario actual: {current_user.Nombre}")
    print(f"ID del usuario: {current_user.Id}")
    print(f"Rol ID: {current_user.RolId}")
    print(f"Rol objeto: {current_user.rol}")
    print(f"Rol nombre: {current_user.rol.Nombre if current_user.rol else 'None'}")
    
    # Verificar si es administrador (aceptar tanto 'admin' como 'Administrador')
    es_admin = current_user.rol and current_user.rol.Nombre.lower() in ['admin', 'administrador']
    
    if not es_admin:
        print(f"Acceso denegado - Rol: {current_user.rol.Nombre if current_user.rol else 'None'}")
        flash('Acceso denegado. Solo los administradores pueden acceder a esta función.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    print("Acceso permitido - Usuario es administrador")
    
    # Obtener todos los posts (activos e inactivos)
    posts = Post.query.order_by(Post.FechaCreacion.desc()).all()
    
    return render_template('admin/admin_gestionar_posts.html', posts=posts)

@admin_bp.route('/admin_estadisticas_foro')
@login_required
@admin_required
def admin_estadisticas_foro():
    """Vista de estadísticas del foro para administradores"""
    # Debug: Imprimir información del usuario actual
    print(f"Usuario actual: {current_user.Nombre}")
    print(f"ID del usuario: {current_user.Id}")
    print(f"Rol ID: {current_user.RolId}")
    print(f"Rol objeto: {current_user.rol}")
    print(f"Rol nombre: {current_user.rol.Nombre if current_user.rol else 'None'}")
    
    # Verificar si es administrador (aceptar tanto 'admin' como 'Administrador')
    es_admin = current_user.rol and current_user.rol.Nombre.lower() in ['admin', 'administrador']
    
    if not es_admin:
        print(f"Acceso denegado - Rol: {current_user.rol.Nombre if current_user.rol else 'None'}")
        flash('Acceso denegado. Solo los administradores pueden acceder a esta función.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    print("Acceso permitido - Usuario es administrador")
    
    # Estadísticas generales
    total_posts = Post.query.count()
    posts_activos = Post.query.filter_by(Estado='Activo').count()
    posts_ocultos = Post.query.filter_by(Estado='Oculto').count()
    posts_eliminados = Post.query.filter_by(Estado='Eliminado').count()
    
    total_comentarios = ComentarioForo.query.count()
    comentarios_activos = ComentarioForo.query.filter_by(Estado='Activo').count()
    comentarios_ocultos = ComentarioForo.query.filter_by(Estado='Oculto').count()
    comentarios_eliminados = ComentarioForo.query.filter_by(Estado='Eliminado').count()
    
    total_likes = Like.query.count()
    
    # Posts más populares
    posts_populares = db.session.query(Post, func.count(Like.Id).label('total_likes')).\
        outerjoin(Like).\
        filter(Post.Estado == 'Activo').\
        group_by(Post.Id).\
        order_by(func.count(Like.Id).desc()).\
        limit(5).all()
    
    # Usuarios más activos
    usuarios_activos = db.session.query(Usuario, func.count(Post.Id).label('total_posts')).\
        outerjoin(Post).\
        filter(Usuario.Estado == 'Activo').\
        group_by(Usuario.Id).\
        order_by(func.count(Post.Id).desc()).\
        limit(5).all()
    
    # Si la petición es AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('format') == 'json':
        return jsonify({
            'total_posts': total_posts,
            'posts_activos': posts_activos,
            'posts_ocultos': posts_ocultos,
            'posts_eliminados': posts_eliminados,
            'total_comentarios': total_comentarios,
            'comentarios_activos': comentarios_activos,
            'comentarios_ocultos': comentarios_ocultos,
            'comentarios_eliminados': comentarios_eliminados,
            'total_likes': total_likes
        })
    
    return render_template('admin/admin_estadisticas_foro.html',
                         total_posts=total_posts,
                         posts_activos=posts_activos,
                         posts_ocultos=posts_ocultos,
                         posts_eliminados=posts_eliminados,
                         total_comentarios=total_comentarios,
                         comentarios_activos=comentarios_activos,
                         comentarios_ocultos=comentarios_ocultos,
                         comentarios_eliminados=comentarios_eliminados,
                         total_likes=total_likes,
                         posts_populares=posts_populares,
                         usuarios_activos=usuarios_activos)

@admin_bp.route('/debug_usuario_foro')
@login_required
@admin_required
def debug_usuario_foro():
    """Ruta de debug para verificar el rol del usuario actual"""
    # Verificar si es administrador (aceptar tanto 'admin' como 'Administrador')
    es_admin = current_user.rol and current_user.rol.Nombre.lower() in ['admin', 'administrador']
    
    usuario_info = {
        'id': current_user.Id,
        'nombre': current_user.Nombre,
        'email': current_user.Email,
        'rol_id': current_user.RolId,
        'rol_nombre': current_user.rol.Nombre if current_user.rol else 'None',
        'estado': current_user.Estado,
        'es_admin': es_admin
    }
    
    return jsonify(usuario_info)

@admin_bp.route('/ayuda')
@login_required
@admin_required
def ayuda():
    """Página de ayuda para administradores"""
    return render_template('admin/ayuda.html')

@admin_bp.route('/manual-admin-pdf')
@login_required
@admin_required
def manual_admin_pdf():
    """Descargar manual de administrador en PDF"""
    try:
        # Ruta al archivo PDF del manual de administrador
        pdf_path = os.path.join(current_app.root_path, 'static', 'docs', 'manual_usuario_admin.pdf')
        
        if os.path.exists(pdf_path):
            return send_file(
                pdf_path,
                as_attachment=True,
                download_name='Manual_Administrador_Flash_Reserver.pdf',
                mimetype='application/pdf'
            )
        else:
            flash('El manual de administrador no está disponible actualmente', 'warning')
            return redirect(url_for('admin.dashboard'))
            
    except Exception as e:
        flash('Error al descargar el manual', 'error')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/actualizar_perfil', methods=['POST'])
@login_required
@admin_required
def actualizar_perfil():
    """Actualizar perfil del administrador"""
    try:
        # Obtener datos del formulario
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        telefono = request.form.get('telefono')
        fecha_nacimiento = request.form.get('fecha_nacimiento')
        direccion = request.form.get('direccion')
        
        # Verificar si el email ya existe en otro usuario
        usuario_existente = Usuario.query.filter(
            Usuario.Email == email,
            Usuario.Id != current_user.Id
        ).first()
        
        if usuario_existente:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'El email ya está registrado por otro usuario'})
            flash('El email ya está registrado por otro usuario', 'error')
            return redirect(url_for('admin.configuracion'))
        
        # Procesar imagen de perfil si se subió
        image_url = None
        if 'foto_perfil' in request.files:
            foto = request.files['foto_perfil']
            if foto and foto.filename:
                # Validar tipo de archivo
                if foto.content_type.startswith('image/'):
                    # Generar nombre seguro
                    filename = secure_filename(foto.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"admin_{current_user.Id}_{timestamp}_{filename}"
                    
                    # Guardar archivo
                    upload_path = os.path.join(current_app.root_path, 'static', 'uploads')
                    os.makedirs(upload_path, exist_ok=True)
                    foto_path = os.path.join(upload_path, filename)
                    foto.save(foto_path)
                    
                    # Actualizar ruta en la base de datos
                    current_user.FotoPerfil = f'uploads/{filename}'
                    image_url = url_for('static', filename=current_user.FotoPerfil)
        
        # Actualizar datos del usuario
        current_user.Nombre = nombre
        current_user.Email = email
        current_user.Telefono = telefono
        current_user.Direccion = direccion
        
        if fecha_nacimiento:
            current_user.FechaNacimiento = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
        
        db.session.commit()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': 'Perfil actualizado exitosamente',
                'image_url': image_url
            })
        
        flash('Perfil actualizado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error al actualizar perfil: {str(e)}'})
        flash(f'Error al actualizar perfil: {str(e)}', 'error')
    
    return redirect(url_for('admin.configuracion'))

@admin_bp.route('/actualizar_configuracion_general', methods=['POST'])
@login_required
@admin_required
def actualizar_configuracion_general():
    """Actualizar configuración general del sistema"""
    try:
        # Obtener datos del formulario
        sitio_nombre = request.form.get('sitio_nombre')
        contacto_email = request.form.get('contacto_email')
        horario_apertura = request.form.get('horario_apertura')
        horario_cierre = request.form.get('horario_cierre')
        zona_horaria = request.form.get('zona_horaria')
        moneda = request.form.get('moneda')
        descripcion_sitio = request.form.get('descripcion_sitio')
        
        # Aquí se guardarían en la base de datos o archivo de configuración
        # Por ahora solo mostramos un mensaje de éxito
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Configuración general actualizada exitosamente'})
        
        flash('Configuración general actualizada exitosamente', 'success')
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error al actualizar configuración: {str(e)}'})
        flash(f'Error al actualizar configuración: {str(e)}', 'error')
    
    return redirect(url_for('admin.configuracion'))

@admin_bp.route('/actualizar_configuracion_reservas', methods=['POST'])
@login_required
@admin_required
def actualizar_configuracion_reservas():
    """Actualizar configuración de reservas del sistema"""
    try:
        # Obtener datos del formulario
        duracion_reserva = request.form.get('duracion_reserva')
        anticipacion_reserva = request.form.get('anticipacion_reserva')
        cancelacion_reserva = request.form.get('cancelacion_reserva')
        
        # Aquí se guardarían en la base de datos o archivo de configuración
        # Por ahora solo mostramos un mensaje de éxito
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Configuración de reservas actualizada exitosamente'})
        
        flash('Configuración de reservas actualizada exitosamente', 'success')
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error al actualizar configuración: {str(e)}'})
        flash(f'Error al actualizar configuración: {str(e)}', 'error')
    
    return redirect(url_for('admin.configuracion'))

@admin_bp.route('/actualizar_configuracion_notificaciones', methods=['POST'])
@login_required
@admin_required
def actualizar_configuracion_notificaciones():
    """Actualizar configuración de notificaciones del sistema"""
    try:
        # Obtener datos del formulario
        email_notificaciones = request.form.get('email_notificaciones')
        push_notificaciones = request.form.get('push_notificaciones')
        
        # Aquí se guardarían en la base de datos o archivo de configuración
        # Por ahora solo mostramos un mensaje de éxito
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Configuración de notificaciones actualizada exitosamente'})
        
        flash('Configuración de notificaciones actualizada exitosamente', 'success')
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error al actualizar configuración: {str(e)}'})
        flash(f'Error al actualizar configuración: {str(e)}', 'error')
    
    return redirect(url_for('admin.configuracion'))

@admin_bp.route('/configuracion/actualizar_campo', methods=['POST'])
@login_required
@admin_required
def actualizar_campo_configuracion():
    """Actualizar un campo específico de configuración"""
    try:
        # Obtener el nombre y valor del campo
        field_name = None
        field_value = None
        
        for key, value in request.form.items():
            field_name = key
            field_value = value
            break
        
        if not field_name:
            return jsonify({'success': False, 'message': 'No se especificó ningún campo'})
        
        # Aquí se guardaría el campo específico en la base de datos o archivo de configuración
        # Por ahora solo mostramos un mensaje de éxito
        
        return jsonify({'success': True, 'message': f'Campo {field_name} actualizado correctamente'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al actualizar campo: {str(e)}'})

@admin_bp.route('/perfil/obtener_info', methods=['GET'])
@login_required
@admin_required
def obtener_info_perfil():
    """Obtener información del perfil del administrador actual"""
    try:
        user_info = {
            'id': current_user.Id,
            'nombre': current_user.Nombre,
            'email': current_user.Email,
            'telefono': current_user.Telefono,
            'direccion': current_user.Direccion,
            'fecha_nacimiento': current_user.FechaNacimiento.isoformat() if current_user.FechaNacimiento else None,
            'imagen': current_user.FotoPerfil,
            'imagen_url': url_for('static', filename=current_user.FotoPerfil) if current_user.FotoPerfil else url_for('static', filename='images/default-user.png')
        }
        
        return jsonify({'success': True, 'data': user_info})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al obtener información: {str(e)}'})

@admin_bp.route('/eliminar_cuenta', methods=['POST'])
@login_required
@admin_required
def eliminar_cuenta():
    """Eliminar cuenta de administrador"""
    try:
        # Verificar que no sea el único administrador
        total_admins = Usuario.query.join(Rol).filter(Rol.Nombre == 'Administrador').count()
        
        if total_admins <= 1:
            flash('No puedes eliminar tu cuenta ya que eres el único administrador del sistema', 'error')
            return redirect(url_for('admin.configuracion'))
        
        # Eliminar usuario
        db.session.delete(current_user)
        db.session.commit()
        
        flash('Tu cuenta ha sido eliminada exitosamente', 'success')
        return redirect(url_for('auth.logout'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar cuenta: {str(e)}', 'error')
        return redirect(url_for('admin.configuracion'))